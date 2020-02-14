import openpyxl
import re
from CMtools import *


class Analysisxls:

    # 此处定义各种单元格处理函数~
    # 简化函数
    def simplepath(self, str_path):

        # 获得如下类型输出('423-太安开关站-6', '422-黄坪变-12'), ('422-黄坪变-14', '305-仁和开关站-7')

        pattern = r"\[(.*?)\-\d{1,2}.*?\-\>(.*?)\-\d{1,2}"

        src = str_path
        match = re.finditer(pattern, src)
        list_re = []
        try:
            for i in match:
                # print(i.group(1))
                list_re.append(i.group(1))
            list_re.append(i.group(2))
        finally:

            return list_re

    def simplepath2(self, str_path):
        # 获得如下类型输出('423-太安开关站-6', '422-黄坪变-12'), ('422-黄坪变-14', '305-仁和开关站-7')

        pattern = r"\[(.*?)\-[N|s].*\-\>(.*?)\-[N|S]"

        src = str_path
        match = re.finditer(pattern, src)
        list_re = []
        for i in match:
            list_re.append(i.groups())

        return list_re

    # 判断是否经过地调
    def paththrughdidiao(self, list_path):

        list_path.append("new col")
        return list_path

    def __init__(self, **kwr):

        self.workbookname = "实时电路可靠性分析工具V039.xlsm"
        self.sheetname = "全路由"
        self.myinit(**kwr)
        self.wb = openpyxl.load_workbook(self.workbookname)
        self.st = self.wb.get_sheet_by_name(self.sheetname)

        # 通过构造字典实现对不同字段列进行函数转换或者新增利用add字段
        self.dic_celltrans = {"8": self.simplepath}
        self.dic_addcell = {"adddidiao": self.paththrughdidiao}
        self.dic_top = {}
        self.dic_top["cha"] = self.dic_celltrans
        self.dic_top["add"] = self.dic_addcell

    def filter(self, **kwr):
        cur = self.st.iter_rows()
        currow = next(cur)
        list_re = []
        # 获取**kwr中的组成一个按顺序排列的字符串

        while currow != self.mainrowtocelltup2():
            currow = next(cur)

        list_tname = []

        for i in currow:
            list_tname.append(i.value)

        list_pd = self.listpd(list_tname, **kwr)

        while currow:

            list_cur = []
            for i in currow:
                list_cur.append(i.value)

            if self.panjue(list_pd, list_cur):
                list_re.append(list_cur)
            try:
                currow = next(cur)
            except:

                return list_re
                # currow = next(cur)

    def panjue(self, p1, s1):

        for p in p1:
            if p == 1:
                continue
            else:
                if len(re.findall(p, s1[p1.index(p)])) == 0:
                    return False

        return True

    def alsomecol(self, list_srt, *licow, **kwr):
        # 使用kwr对应的方法来控制函数执行的过滤。譬如path="simple"路径简化
        def chuli1(n, **licow1):  #
            #
            for j in licow1["cha"]:
                n[int(j)] = licow1["cha"][j](n[int(j)])
            for j in licow1["add"]:
                n = licow1["add"][j](n)
            return n

        for i in list_srt:
            yield chuli1(i, **kwr)

    def list_getsheetcontent(self, rowrange):
        list_r = []

        # list_row=rowrange.split(":")
        resul = self.st[rowrange]
        for i in resul:
            temlist = []
            for j in i:
                temlist.append(j.value)
            list_r.append(temlist)
        return list_r

    def listpd(self, list_tname, **kwr):
        list_p = len(list_tname) * [1]
        for key in kwr:
            iter_li = iter(list_tname)
            plist = next(iter_li)  # type:str
            while plist:
                if plist.find(key) > -1:
                    list_p[list_tname.index(plist)] = kwr[key]
                    break
                plist = next(iter_li)

        return list_p

    def myinit(self, **kwr):
        for a in kwr:
            if a == "workbook":
                self.workbookname = kwr[a]
            elif a == "sheet":
                self.sheetname = kwr[a]

    def __str__(self):
        self.filter()
        return str(self.mainrowtocelltup2())

    def mainrowtocelltup2(self):
        # 根据表格中第一行等于整体表格宽度来判断是否是表头别，利用生成器实现

        ll = self.st.max_column

        # print (type(st._cells))
        num = 1
        perrow = self.st.iter_rows()
        while True:
            exit_flag = False

            currow = next(perrow)
            for i in currow:

                if i.value == None:
                    exit_flag = True
                    break

            if exit_flag:
                continue

            return currow

    def list_getsheettitile(self):
        # 生成表头的列表
        list_re = []

        ll = self.st.max_column

        # print (type(st._cells))
        num = 1
        perrow = self.st.iter_rows()
        while True:
            exit_flag = False

            currow = next(perrow)
            for i in currow:

                if i.value == None:
                    exit_flag = True
                    break

            if exit_flag:
                continue
            for i in currow:
                list_re.append(i.value)
            return list_re

    def process(self):
        pass


def main(argv=None):
    # SDH[101-省调-14-SL16A-1(SDH-1)-VC4:4->105-厂口变-9-SL16A-1(SDH-1)-VC4:4]
    dicc = {"方向": "双向", "工作路由": "版纳地调.*-6.*->"}
    c = Analysisxls(**dicc)

    # list_text = c.list_getsheetcontent("19:20")
    # print("test:", list_text[1][8])
    #
    # pattern = "dfdfdf"
    # match = re.findall(pattern, list_text[1][8])
    # print(len(match))

    lista = c.filter(**dicc)
    sum=0
    for i in c.alsomecol(lista,**c.dic_top):
        print(i)

    print(sum)



if __name__ == "__main__":
    main()
