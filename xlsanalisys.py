#引入工作表 同时对字段进行筛选。
import openpyxl
import re

def mainrowtocelltup(book,sheet):
    #根据表格中第一行等于整体表格宽度来判断是否是表头别,利用forx循环
    wb = openpyxl.load_workbook(book)
    st = wb.get_sheet_by_name(sheet)

    ll=st.max_column

    #print (type(st._cells))
    num=1

    for perrow in st.iter_rows():
        exit_flag=False

        for pcell in perrow:


            if pcell.value ==1:
                print("we get 1")
                print(type(pcell.value))
                exit_flag=True

                break

        if exit_flag:
            continue
        else:
            break

    return perrow

def mainrowtocelltup2(book,sheet):
    #根据表格中第一行等于整体表格宽度来判断是否是表头别，利用生成器实现

    ll=st.max_column

    #print (type(st._cells))
    num=1
    perrow = st.iter_rows()
    while True:
        exit_flag=False

        currow=next(perrow)
        for i in currow:

            if i.value==None:
                exit_flag=True
                break

        if exit_flag:
            continue

        return currow
#################使用while循环实现
    # while True:
    #     exit_flag = False
    #     # tup_tittle=  next(st.iter_rows())
    #     # tup_tittle = next(st.iter_rows())
    #     tup_tittle=iter(st.iter_rows())
    #     tup_tittle=st.iter_rows().__next__()
    #     tup_tittle2 = st.iter_rows().__next__()
    #
    #
    #     for i in range(1,3):
    #         print(tup_tittle.__next__())
    #
    #
    #        #print(st.iter_rows().__next__())
    #        #print(st.iter_rows().__next__())
    #
    #
    #
    #
    #     print(type(st.iter_rows()))
    #     print(type(tup_tittle))
    #     print("tup is {}".format(tup_tittle))
    #     print("tup2 is {}".format(tup_tittle2))
    #
    #     if len(tup_tittle)==ll:
    #         for i in tup_tittle:
    #             if i==1:
    #                 exit_flag = True
    #                 break
    #         if exit_flag:
    #             continue
    #         else:
    #             break







    # for a in st.iter_rows():
    #     #print (a,type(a),type(st._cells[a]))
    #     print(a,len(a))
    #     num+=1
    #     if num>5:
    #         break
    # while a!=none:
    #    lis=a
    #    # if len(a)=1:
    #       #将该列表转化为一个fdic
    # fdic={"kunming":"liugan"}

workbookname="实时电路可靠性分析工具V039.xlsm"
sheetname="全路由"
#定义
wb=openpyxl.load_workbook(workbookname)
st=wb.get_sheet_by_name(sheetname)#
st_col_mpath=st["b"]
#print (len(st_col_mpath))

my_dic=mainrowtocelltup2(workbookname,sheetname)
iterrow=st.iter_rows()
# for i in range(0,3):
#     print(next(iterrow),type(iterrow))
num2=1
while True:
    print(next(iterrow), type(iterrow))
    num2+=1
    if num2>3:
        break



#func 传入一个字典，可以对指定列数进行筛选，用生成器来实现,输出为一个列表
def analasyskey(workbook,sheetname,**kwargs):
    workbookname = "实时电路可靠性分析工具V039.xlsm"
    sheetname = "全路由"
    tittle_row=mainrowtocelltup2(workbookname,sheetname)
    result=[]

    st=wb.get_sheet_by_name(sheetname)
    currow=st.iter_rows()
    prow=next(currow)
    i=0
    while  prow != tittle_row:
        print("prow is :{}".format(prow))
        print("tittle_row{}".format(tittle_row))
        prow=next(currow)
        i+=1

    print("title is :{}".format(prow))


    num=1
    while prow:
  #      if 1:
            #filcur=filtercur(curow)
            filcur=prow
            result.append(filcur)
            print(prow)
            num+=1

            if num>4:
                break
            prow = next(currow)


    return result

print("func output is :")
print(len(analasyskey(workbookname, sheetname)))

